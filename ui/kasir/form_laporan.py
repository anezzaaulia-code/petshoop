# -*- coding: utf-8 -*-
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import (
    QTableWidgetItem,
    QHeaderView,
    QFileDialog,
    QMessageBox
)
from database.connect import get_connection

class Ui_FormLaporan(object):

    # =========================================================
    #                       SETUP UI
    # =========================================================
    def setupUi(self, MainWindow, kasir_login=""):
        self.MainWindow = MainWindow
        self.kasir_login = kasir_login  

        MainWindow.setObjectName("FormLaporan")
        MainWindow.resize(1150, 800)

        # ====== GLOBAL STYLE ======
        # Catatan: cursor: pointer dihapus dari CSS agar terminal bersih
        MainWindow.setStyleSheet("""
            QWidget {
                font-family: 'Segoe UI', sans-serif;
                font-size: 14px;
                color: #2c3e50;
                background-color: #f4f6f9;
            }
            
            /* Card Style */
            QFrame#FilterCard, QFrame#TableCard, QFrame#TotalCard {
                background-color: #ffffff;
                border-radius: 10px;
                border: 1px solid #e0e0e0;
            }

            /* Input Style */
            QLineEdit, QDateEdit {
                padding: 8px 12px;
                border: 1px solid #ccc;
                border-radius: 6px;
                background-color: #fff;
            }
            QLineEdit:focus, QDateEdit:focus {
                border: 1px solid #3498db;
            }

            /* Button Style */
            QPushButton {
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: 600;
                border: none;
            }
            QPushButton:hover {
                opacity: 0.9;
            }

            QPushButton#btnFilter { background-color: #3498db; color: white; }
            QPushButton#btnReset { background-color: #95a5a6; color: white; }
            QPushButton#btnExport { background-color: #27ae60; color: white; }
            
            QPushButton#btnBack {
                background-color: transparent;
                color: #7f8c8d;
                border: 1px solid #bdc3c7;
            }
            QPushButton#btnBack:hover {
                background-color: #ecf0f1;
                color: #2c3e50;
            }

            /* Table Style */
            QTableWidget {
                border: none;
                gridline-color: #f0f0f0;
                background-color: white;
            }
            QHeaderView::section {
                background-color: #f8f9fa;
                padding: 10px;
                border: none;
                border-bottom: 2px solid #ddd;
                font-weight: bold;
                color: #555;
            }
            QTableWidget::item:selected {
                background-color: #e8f6ff;
                color: #2980b9;
            }
        """)

        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.mainLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.mainLayout.setContentsMargins(25, 25, 25, 25)
        self.mainLayout.setSpacing(20)

        # ===== HEADER =====
        self.headerLayout = QtWidgets.QHBoxLayout()
        self.titleLabel = QtWidgets.QLabel("📄 Laporan Transaksi")
        self.titleLabel.setStyleSheet("font-size: 24px; font-weight: bold; color: #2c3e50;")
        self.headerLayout.addWidget(self.titleLabel)
        self.headerLayout.addStretch()
        self.mainLayout.addLayout(self.headerLayout)

        # ===== FILTER CARD =====
        self.filterCard = QtWidgets.QFrame()
        self.filterCard.setObjectName("FilterCard")
        self.filterLayout = QtWidgets.QHBoxLayout(self.filterCard)
        self.filterLayout.setContentsMargins(15, 15, 15, 15)
        self.filterLayout.setSpacing(10)

        # Date Inputs
        self.labelDate = QtWidgets.QLabel("📅 Periode:")
        self.labelDate.setStyleSheet("font-weight: bold;")
        
        self.dateStart = QtWidgets.QDateEdit()
        self.dateStart.setCalendarPopup(True)
        self.dateStart.setDate(QtCore.QDate.currentDate())
        self.dateStart.setDisplayFormat("dd MMM yyyy")
        self.dateStart.setFixedWidth(130)

        self.labelTo = QtWidgets.QLabel("- s/d -")
        
        self.dateEnd = QtWidgets.QDateEdit()
        self.dateEnd.setCalendarPopup(True)
        self.dateEnd.setDate(QtCore.QDate.currentDate())
        self.dateEnd.setDisplayFormat("dd MMM yyyy")
        self.dateEnd.setFixedWidth(130)

        # Kasir Input
        self.inputKasir = QtWidgets.QLineEdit()
        self.inputKasir.setPlaceholderText("👤 Cari nama kasir...")
        self.inputKasir.setFixedWidth(200)
        if self.kasir_login:
            self.inputKasir.setText(self.kasir_login)

        # Buttons (Set Cursor via Code)
        self.btnFilter = QtWidgets.QPushButton("🔍 Filter")
        self.btnFilter.setObjectName("btnFilter")
        self.btnFilter.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        self.btnReset = QtWidgets.QPushButton("🔄 Reset")
        self.btnReset.setObjectName("btnReset")
        self.btnReset.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        self.btnExport = QtWidgets.QPushButton("📥 Export Excel")
        self.btnExport.setObjectName("btnExport")
        self.btnExport.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        self.filterLayout.addWidget(self.labelDate)
        self.filterLayout.addWidget(self.dateStart)
        self.filterLayout.addWidget(self.labelTo)
        self.filterLayout.addWidget(self.dateEnd)
        self.filterLayout.addWidget(self.inputKasir)
        self.filterLayout.addWidget(self.btnFilter)
        self.filterLayout.addWidget(self.btnReset)
        self.filterLayout.addStretch()
        self.filterLayout.addWidget(self.btnExport)

        self.mainLayout.addWidget(self.filterCard)

        # ===== TABLE CARD =====
        self.tableCard = QtWidgets.QFrame()
        self.tableCard.setObjectName("TableCard")
        self.tableLayout = QtWidgets.QVBoxLayout(self.tableCard)
        self.tableLayout.setContentsMargins(0, 0, 0, 0)

        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(
            ["ID TRX", "Tanggal & Waktu", "Nama Kasir", "Total Transaksi", "Detail Produk"]
        )

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setFocusPolicy(QtCore.Qt.NoFocus)

        self.tableLayout.addWidget(self.table)
        self.mainLayout.addWidget(self.tableCard)

        # ===== FOOTER =====
        self.footerLayout = QtWidgets.QHBoxLayout()

        self.btnBack = QtWidgets.QPushButton("⬅ Kembali")
        self.btnBack.setObjectName("btnBack")
        self.btnBack.setFixedWidth(200)
        self.btnBack.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))

        self.totalCard = QtWidgets.QFrame()
        self.totalCard.setObjectName("TotalCard")
        self.totalCard.setStyleSheet("background-color: #2c3e50; color: white; border: none;")
        self.totalLayout = QtWidgets.QHBoxLayout(self.totalCard)
        self.totalLayout.setContentsMargins(20, 10, 20, 10)

        self.lblTotalText = QtWidgets.QLabel("Total Pemasukan:")
        self.lblTotalText.setStyleSheet("font-weight: normal; font-size: 14px; color: #ecf0f1;")

        self.lblTotalValue = QtWidgets.QLabel("Rp 0")
        self.lblTotalValue.setStyleSheet("font-size: 20px; font-weight: bold; color: #2ecc71;")

        self.totalLayout.addWidget(self.lblTotalText)
        self.totalLayout.addSpacing(15)
        self.totalLayout.addWidget(self.lblTotalValue)

        self.footerLayout.addWidget(self.btnBack)
        self.footerLayout.addStretch()
        self.footerLayout.addWidget(self.totalCard)

        self.mainLayout.addLayout(self.footerLayout)
        MainWindow.setCentralWidget(self.centralwidget)

        # ===== EVENTS =====
        self.btnFilter.clicked.connect(self.filterData)
        self.btnReset.clicked.connect(self.loadAllData)
        self.btnBack.clicked.connect(self.goBack)
        self.btnExport.clicked.connect(self.exportToExcel)

        # Auto Load
        QtCore.QTimer.singleShot(100, self.loadAllData)

    # ================= LOGIC LOAD DATA =================
    def loadAllData(self):
        try:
            db, cursor = get_connection()
            if not db: return 

            if self.kasir_login: 
                cursor.execute("""
                    SELECT id_transaksi, tanggal, kasir, total 
                    FROM transaksi 
                    WHERE kasir=%s
                    ORDER BY tanggal DESC
                """, (self.kasir_login,))
            else:
                cursor.execute("SELECT id_transaksi, tanggal, kasir, total FROM transaksi ORDER BY tanggal DESC")

            rows = cursor.fetchall()
            cursor.close()
            db.close()
            self.displayData(rows)
            
            # Reset visual
            self.inputKasir.setText(self.kasir_login if self.kasir_login else "")
            self.dateStart.setDate(QtCore.QDate.currentDate())
            self.dateEnd.setDate(QtCore.QDate.currentDate())

        except Exception as e:
            print(f"Error loading data: {e}")

    def filterData(self):
        start = self.dateStart.date().toString("yyyy-MM-dd")
        end = self.dateEnd.date().toString("yyyy-MM-dd")
        kasir = self.inputKasir.text()

        query = """
            SELECT id_transaksi, tanggal, kasir, total
            FROM transaksi
            WHERE DATE(tanggal) BETWEEN %s AND %s
        """
        params = [start, end]

        if kasir:
            query += " AND kasir LIKE %s"
            params.append(f"%{kasir}%")

        query += " ORDER BY tanggal DESC"

        try:
            db, cursor = get_connection()
            if not db: return
            cursor.execute(query, params)
            rows = cursor.fetchall()
            cursor.close()
            db.close()
            self.displayData(rows)
        except Exception as e:
            print(f"Error filtering data: {e}")

    def displayData(self, rows):
        self.table.setRowCount(0)
        self.table.setRowCount(len(rows))
        total_pemasukan = 0

        for i, row in enumerate(rows):
            id_trx, tanggal, kasir, total = row
            produk_list = self.getProdukList(id_trx)

            item_id = QTableWidgetItem(str(id_trx))
            item_id.setTextAlignment(QtCore.Qt.AlignCenter)
            
            item_tgl = QTableWidgetItem(str(tanggal))
            
            item_kasir = QTableWidgetItem(str(kasir))
            item_kasir.setTextAlignment(QtCore.Qt.AlignCenter)

            item_total = QTableWidgetItem(f"Rp {total:,}".replace(",", "."))
            item_total.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Bold))

            self.table.setItem(i, 0, item_id)
            self.table.setItem(i, 1, item_tgl)
            self.table.setItem(i, 2, item_kasir)
            self.table.setItem(i, 3, item_total)
            self.table.setItem(i, 4, QTableWidgetItem(produk_list))

            total_pemasukan += total

        self.lblTotalValue.setText(f"Rp {total_pemasukan:,}".replace(",", "."))

    def getProdukList(self, id_transaksi):
        try:
            db, cursor = get_connection()
            cursor.execute("""
                SELECT produk.nama_produk, detail_transaksi.qty
                FROM detail_transaksi
                JOIN produk ON produk.id_produk = detail_transaksi.id_produk
                WHERE detail_transaksi.id_transaksi = %s
            """, (id_transaksi,))
            rows = cursor.fetchall()
            cursor.close()
            db.close()
            if not rows: return "-"
            return ", ".join([f"{nama} ({qty})" for nama, qty in rows])
        except:
            return "Error load detail"

    def exportToExcel(self):
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.warning(None, "Error Library", "Library 'openpyxl' belum terinstal.\nSilakan install: pip install openpyxl")
            return
        
        if self.table.rowCount() == 0:
            QMessageBox.warning(None, "Data Kosong", "Tidak ada data untuk diexport.")
            return

        path, _ = QFileDialog.getSaveFileName(None, "Simpan Laporan Excel", f"Laporan_Transaksi_{QtCore.QDate.currentDate().toString('yyyyMMdd')}.xlsx", "Excel Files (*.xlsx)")
        if not path: return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Laporan"
            headers = ["ID", "Tanggal", "Kasir", "Total (Rp)", "Produk Dibeli"]
            ws.append(headers)

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for col, h in enumerate(headers, 1):
                c = ws.cell(1, col)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = openpyxl.styles.PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border

            for r in range(self.table.rowCount()):
                row_data = []
                for c in range(self.table.columnCount()):
                    text = self.table.item(r, c).text()
                    if c == 3: 
                        clean = text.replace("Rp", "").replace(".", "").strip()
                        row_data.append(int(clean) if clean.isdigit() else 0)
                    else:
                        row_data.append(text)
                ws.append(row_data)
                for c_idx in range(1, len(headers) + 1):
                    ws.cell(r + 2, c_idx).border = thin_border

            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

            wb.save(path)
            QMessageBox.information(None, "Sukses", f"Laporan berhasil disimpan di:\n{path}")
        except Exception as e:
            QMessageBox.critical(None, "Error Export", f"Gagal export data: {e}")

    def goBack(self):
        try:
            from dashboard.dashboard_kasir import Ui_MainWindow as DashboardKasir
            self.window_dashboard = QtWidgets.QMainWindow()
            self.ui_dashboard = DashboardKasir()
            try:
                self.ui_dashboard.setupUi(self.window_dashboard, self.kasir_login)
            except TypeError:
                self.ui_dashboard.setupUi(self.window_dashboard)
            self.window_dashboard.show()
            self.MainWindow.close()
        except Exception as e:
             QMessageBox.critical(None, "Error", f"Gagal kembali ke dashboard: {e}")

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    win = QtWidgets.QMainWindow()
    ui = Ui_FormLaporan()
    ui.setupUi(win, kasir_login="Admin")
    win.show()
    sys.exit(app.exec_())